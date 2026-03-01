"""Tests for xldump.extractors module."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from xldump.extractors import (
    extract_cells,
    extract_images,
    extract_merged_cells,
    extract_sheet_info,
    extract_validations,
)
from xldump.extractors.color import (
    DEFAULT_THEME_COLORS,
    INDEXED_COLORS,
    _apply_tint,
    get_color_cache_key,
    resolve_color,
)
from xldump.extractors.image import has_images
from xldump.extractors.validation import has_data_validations


class TestExtractCells:
    """Tests for extract_cells() function."""

    def test_extract_simple_cells(self, simple_table_path: Path) -> None:
        """Test extracting cells from simple table."""
        wb = load_workbook(simple_table_path)
        ws = wb.active
        assert ws is not None

        rows = extract_cells(ws)

        # Check that rows were extracted
        assert len(rows) > 0

        # Check header row
        assert "1" in rows
        row1 = rows["1"]
        assert "A" in row1
        assert row1["A"] is not None
        assert row1["A"].value == "ID"

        wb.close()

    def test_extract_cells_with_styles(self, styled_cells_path: Path) -> None:
        """Test extracting cells with style information."""
        wb = load_workbook(styled_cells_path)
        ws = wb.active
        assert ws is not None

        rows = extract_cells(ws, include_styles=True)

        # Check cell A1 has bold font
        assert "1" in rows
        assert "A" in rows["1"]
        cell = rows["1"]["A"]
        assert cell is not None
        assert cell.style is not None
        assert cell.style.font is not None
        assert cell.style.font.bold is True

        wb.close()

    def test_extract_cells_no_styles(self, styled_cells_path: Path) -> None:
        """Test extracting cells without style information."""
        wb = load_workbook(styled_cells_path)
        ws = wb.active
        assert ws is not None

        rows = extract_cells(ws, include_styles=False)

        # Check that style is None
        for row in rows.values():
            for cell in row.values():
                if cell is not None:
                    assert cell.style is None

        wb.close()

    def test_extract_cells_no_values(self, simple_table_path: Path) -> None:
        """Test extracting cells without values."""
        wb = load_workbook(simple_table_path)
        ws = wb.active
        assert ws is not None

        rows = extract_cells(ws, include_values=False)

        # Check that values are None
        for row in rows.values():
            for cell in row.values():
                if cell is not None:
                    assert cell.value is None

        wb.close()

    def test_extract_merged_cells_content(self, merged_cells_path: Path) -> None:
        """Test that merged cells have correct structure."""
        wb = load_workbook(merged_cells_path)
        ws = wb.active
        assert ws is not None

        rows = extract_cells(ws)

        # The merged range A1:D1 should have value in A1 only
        assert "1" in rows
        assert "A" in rows["1"]

        # A1 should have the value
        a1 = rows["1"]["A"]
        assert a1 is not None
        assert a1.value == "Section Title"

        # B1, C1, D1 should be None (part of merge, not top-left)
        if "B" in rows["1"]:
            assert rows["1"]["B"] is None
        if "C" in rows["1"]:
            assert rows["1"]["C"] is None
        if "D" in rows["1"]:
            assert rows["1"]["D"] is None

        wb.close()


class TestExtractMergedCells:
    """Tests for extract_merged_cells() function."""

    def test_extract_merged_cells(self, merged_cells_path: Path) -> None:
        """Test extracting merged cell ranges."""
        wb = load_workbook(merged_cells_path)
        ws = wb.active
        assert ws is not None

        merged = extract_merged_cells(ws)

        assert len(merged) > 0

        # Check that A1:D1 merge is captured
        ranges = [mc.range for mc in merged]
        assert any("A1" in r and "D1" in r for r in ranges)

        # Check merge info structure
        for mc in merged:
            assert mc.min_row > 0
            assert mc.max_row >= mc.min_row
            assert mc.min_col > 0
            assert mc.max_col >= mc.min_col

        wb.close()

    def test_no_merged_cells(self, simple_table_path: Path) -> None:
        """Test extracting from workbook without merged cells."""
        wb = load_workbook(simple_table_path)
        ws = wb.active
        assert ws is not None

        merged = extract_merged_cells(ws)

        assert len(merged) == 0

        wb.close()


class TestExtractSheetInfo:
    """Tests for extract_sheet_info() function."""

    def test_extract_sheet_info(self, simple_table_path: Path) -> None:
        """Test extracting sheet info."""
        wb = load_workbook(simple_table_path, read_only=False)
        ws = wb.active
        assert ws is not None

        info = extract_sheet_info(ws, 0)

        assert info.name == "Sheet1"
        assert info.index == 0
        assert info.max_row >= 4
        assert info.max_column >= 3

        wb.close()

    def test_extract_sheet_info_with_merges(self, merged_cells_path: Path) -> None:
        """Test sheet info includes merged cell count."""
        wb = load_workbook(merged_cells_path, read_only=False)
        ws = wb.active
        assert ws is not None

        info = extract_sheet_info(ws, 0)

        assert info.merged_cell_count > 0

        wb.close()


class TestExtractValidations:
    """Tests for extract_validations() function."""

    def test_extract_validations(self, data_validations_path: Path) -> None:
        """Test extracting data validations."""
        wb = load_workbook(data_validations_path)
        ws = wb.active
        assert ws is not None

        validations = extract_validations(ws)

        assert len(validations) == 2

        # Check list validation
        list_validation = next((v for v in validations if v.type == "list"), None)
        assert list_validation is not None
        assert '"Active,Inactive,Pending"' in (list_validation.formula1 or "")
        assert list_validation.prompt == "Select a status"
        assert list_validation.error == "Invalid status"

        # Check whole number validation
        whole_validation = next((v for v in validations if v.type == "whole"), None)
        assert whole_validation is not None
        assert whole_validation.formula1 == "1"
        assert whole_validation.formula2 == "100"
        assert whole_validation.allow_blank is False

        wb.close()

    def test_has_data_validations_true(self, data_validations_path: Path) -> None:
        """Test has_data_validations returns True for sheet with validations."""
        wb = load_workbook(data_validations_path)
        ws = wb.active
        assert ws is not None

        assert has_data_validations(ws) is True

        wb.close()

    def test_has_data_validations_false(self, simple_table_path: Path) -> None:
        """Test has_data_validations returns False for sheet without validations."""
        wb = load_workbook(simple_table_path)
        ws = wb.active
        assert ws is not None

        assert has_data_validations(ws) is False

        wb.close()


class TestExtractImages:
    """Tests for extract_images() function."""

    def test_extract_images(self, with_images_path: Path) -> None:
        """Test extracting image information."""
        wb = load_workbook(with_images_path)
        ws = wb.active
        assert ws is not None

        images = extract_images(ws)

        assert len(images) == 1

        img_info = images[0]
        assert img_info.anchor == "B2"
        # Width/height come from actual PNG dimensions (1x1 pixel test image)
        assert img_info.width == 1
        assert img_info.height == 1

        wb.close()

    def test_has_images_true(self, with_images_path: Path) -> None:
        """Test has_images returns True for sheet with images."""
        wb = load_workbook(with_images_path)
        ws = wb.active
        assert ws is not None

        assert has_images(ws) is True

        wb.close()

    def test_has_images_false(self, simple_table_path: Path) -> None:
        """Test has_images returns False for sheet without images."""
        wb = load_workbook(simple_table_path)
        ws = wb.active
        assert ws is not None

        assert has_images(ws) is False

        wb.close()


class TestColorResolution:
    """Tests for color resolution functions."""

    def test_resolve_color_none(self) -> None:
        """Test resolve_color with None input."""
        result = resolve_color(None)
        assert result is None

    def test_resolve_color_rgb_8digit(self) -> None:
        """Test resolve_color with 8-digit RGB color."""
        from openpyxl.styles.colors import Color

        color = Color(rgb="FFFF0000")  # Red with alpha
        result = resolve_color(color)
        assert result == "FFFF0000"

    def test_resolve_color_rgb_6digit(self) -> None:
        """Test resolve_color with 6-digit RGB color.

        Note: openpyxl auto-pads 6-digit RGB to 8-digit with "00" prefix,
        so "00FF00" becomes "0000FF00" internally.
        """
        from openpyxl.styles.colors import Color

        # openpyxl converts "00FF00" to "0000FF00" internally
        color = Color(rgb="00FF00")  # Green
        result = resolve_color(color)
        # Result is 8-digit, already prefixed by openpyxl
        assert result == "0000FF00"

    def test_resolve_color_indexed(self) -> None:
        """Test resolve_color with indexed color."""
        from openpyxl.styles.colors import Color

        # Index 0 is black
        color = Color(indexed=0)
        result = resolve_color(color)
        assert result == "FF000000"

    def test_resolve_color_theme_default(self) -> None:
        """Test resolve_color with theme color using defaults."""
        from openpyxl.styles.colors import Color

        # Theme 0 is dark1 (black in default theme)
        color = Color(theme=0)
        result = resolve_color(color, workbook=None)
        assert result == "FF000000"

    def test_apply_tint_positive(self) -> None:
        """Test _apply_tint with positive tint (lighter)."""
        # Pure black with 0.5 tint should become gray
        result = _apply_tint("000000", 0.5)
        # Check it's lighter (higher luminance)
        r = int(result[0:2], 16)
        g = int(result[2:4], 16)
        b = int(result[4:6], 16)
        assert r > 0 and g > 0 and b > 0  # Should no longer be black

    def test_apply_tint_negative(self) -> None:
        """Test _apply_tint with negative tint (darker)."""
        # Pure white with -0.5 tint should become gray
        result = _apply_tint("FFFFFF", -0.5)
        # Check it's darker (lower luminance)
        r = int(result[0:2], 16)
        g = int(result[2:4], 16)
        b = int(result[4:6], 16)
        assert r < 255 and g < 255 and b < 255  # Should no longer be white

    def test_apply_tint_zero(self) -> None:
        """Test _apply_tint with zero tint (no change)."""
        original = "FF0000"
        result = _apply_tint(original, 0.0)
        # Due to floating point, may not be exact, but should be very close
        r = int(result[0:2], 16)
        assert r >= 254  # Red should stay close to 255

    def test_default_theme_colors(self) -> None:
        """Test DEFAULT_THEME_COLORS has expected entries."""
        assert len(DEFAULT_THEME_COLORS) == 12
        assert DEFAULT_THEME_COLORS[0] == "000000"  # dk1 (dark 1)
        assert DEFAULT_THEME_COLORS[1] == "FFFFFF"  # lt1 (light 1)

    def test_indexed_colors(self) -> None:
        """Test INDEXED_COLORS has expected entries."""
        assert INDEXED_COLORS[0] == "000000"  # Black
        assert INDEXED_COLORS[1] == "FFFFFF"  # White
        assert INDEXED_COLORS[2] == "FF0000"  # Red
        assert INDEXED_COLORS[3] == "00FF00"  # Green
        assert INDEXED_COLORS[4] == "0000FF"  # Blue

    def test_get_color_cache_key_rgb(self) -> None:
        """Test get_color_cache_key for RGB color.

        Note: openpyxl auto-pads 6-digit RGB to 8-digit.
        """
        from openpyxl.styles.colors import Color

        # Use 8-digit ARGB format to match openpyxl's internal representation
        color = Color(rgb="FFFF0000")  # Red with FF alpha
        key = get_color_cache_key(color)
        assert key == "rgb:FFFF0000"

    def test_get_color_cache_key_theme(self) -> None:
        """Test get_color_cache_key for theme color."""
        from openpyxl.styles.colors import Color

        color = Color(theme=1, tint=0.5)
        key = get_color_cache_key(color)
        assert key == "theme:1:0.5"

    def test_get_color_cache_key_indexed(self) -> None:
        """Test get_color_cache_key for indexed color."""
        from openpyxl.styles.colors import Color

        color = Color(indexed=5)
        key = get_color_cache_key(color)
        assert key == "indexed:5"

    def test_get_color_cache_key_none(self) -> None:
        """Test get_color_cache_key with None input."""
        key = get_color_cache_key(None)
        assert key is None
