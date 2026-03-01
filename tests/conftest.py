"""Pytest configuration and fixtures for xldump tests.

This module provides test fixtures including programmatically generated
Excel files for testing various features.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
import pytest

FIXTURES_DIR = Path(__file__).parent / "fixtures"


@pytest.fixture(scope="session", autouse=True)
def create_fixtures() -> None:
    """Create test .xlsx fixtures programmatically."""
    FIXTURES_DIR.mkdir(exist_ok=True)
    _create_simple_table()
    _create_merged_cells()
    _create_styled_cells()
    _create_multi_sheet()
    _create_data_validations()
    _create_with_images()


@pytest.fixture
def fixtures_dir() -> Path:
    """Return the fixtures directory path."""
    return FIXTURES_DIR


@pytest.fixture
def simple_table_path(fixtures_dir: Path) -> Path:
    """Return path to simple_table.xlsx fixture."""
    return fixtures_dir / "simple_table.xlsx"


@pytest.fixture
def merged_cells_path(fixtures_dir: Path) -> Path:
    """Return path to merged_cells.xlsx fixture."""
    return fixtures_dir / "merged_cells.xlsx"


@pytest.fixture
def styled_cells_path(fixtures_dir: Path) -> Path:
    """Return path to styled_cells.xlsx fixture."""
    return fixtures_dir / "styled_cells.xlsx"


@pytest.fixture
def multi_sheet_path(fixtures_dir: Path) -> Path:
    """Return path to multi_sheet.xlsx fixture."""
    return fixtures_dir / "multi_sheet.xlsx"


@pytest.fixture
def data_validations_path(fixtures_dir: Path) -> Path:
    """Return path to data_validations.xlsx fixture."""
    return fixtures_dir / "data_validations.xlsx"


@pytest.fixture
def with_images_path(fixtures_dir: Path) -> Path:
    """Return path to with_images.xlsx fixture."""
    return fixtures_dir / "with_images.xlsx"


def _create_simple_table() -> None:
    """Create a simple table fixture with basic data."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.title = "Sheet1"

    # Header row
    ws.append(["ID", "Name", "Value"])

    # Data rows
    ws.append([1, "Item A", 100])
    ws.append([2, "Item B", 200])
    ws.append([3, "Item C", 300])

    wb.save(FIXTURES_DIR / "simple_table.xlsx")


def _create_merged_cells() -> None:
    """Create a fixture with merged cells."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.title = "Sheet1"

    # Merged header
    ws.merge_cells("A1:D1")
    ws["A1"] = "Section Title"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Sub-header
    ws.merge_cells("A2:B2")
    ws["A2"] = "Category"
    ws.merge_cells("C2:D2")
    ws["C2"] = "Details"

    # Data
    ws["A3"] = "Item 1"
    ws["B3"] = "Description 1"
    ws["C3"] = 100
    ws["D3"] = "Note 1"

    wb.save(FIXTURES_DIR / "merged_cells.xlsx")


def _create_styled_cells() -> None:
    """Create a fixture with various cell styles."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.title = "Styled"

    # Bold and italic
    ws["A1"] = "Bold"
    ws["A1"].font = Font(bold=True)

    ws["B1"] = "Italic"
    ws["B1"].font = Font(italic=True)

    ws["C1"] = "Bold Italic"
    ws["C1"].font = Font(bold=True, italic=True)

    # Colors
    ws["A2"] = "Red Text"
    ws["A2"].font = Font(color="FF0000")

    ws["B2"] = "Blue Background"
    ws["B2"].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

    # Alignment
    ws["A3"] = "Center"
    ws["A3"].alignment = Alignment(horizontal="center")

    ws["B3"] = "Right"
    ws["B3"].alignment = Alignment(horizontal="right")

    ws["C3"] = "Wrapped Text That Should Wrap"
    ws["C3"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["C"].width = 15

    # Borders
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ws["A4"] = "Bordered"
    ws["A4"].border = thin_border

    # Number format
    ws["A5"] = 0.5
    ws["A5"].number_format = "0.00%"

    ws["B5"] = 1234.56
    ws["B5"].number_format = "#,##0.00"

    # Font size
    ws["A6"] = "Large"
    ws["A6"].font = Font(size=18)

    ws["B6"] = "Small"
    ws["B6"].font = Font(size=8)

    wb.save(FIXTURES_DIR / "styled_cells.xlsx")


def _create_multi_sheet() -> None:
    """Create a fixture with multiple sheets."""
    wb = Workbook()

    # First sheet (default)
    ws1 = wb.active
    if ws1 is None:
        return
    ws1.title = "Summary"
    ws1["A1"] = "Summary Sheet"
    ws1["A2"] = "Total"
    ws1["B2"] = 600

    # Second sheet
    ws2 = wb.create_sheet("Data")
    ws2["A1"] = "ID"
    ws2["B1"] = "Value"
    ws2["A2"] = 1
    ws2["B2"] = 100
    ws2["A3"] = 2
    ws2["B3"] = 200
    ws2["A4"] = 3
    ws2["B4"] = 300

    # Third sheet
    ws3 = wb.create_sheet("Notes")
    ws3["A1"] = "Notes"
    ws3["A2"] = "This is a note."

    wb.save(FIXTURES_DIR / "multi_sheet.xlsx")


def _create_data_validations() -> None:
    """Create a fixture with data validations."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.title = "Sheet1"

    # Header row
    ws["A1"] = "Status"
    ws["B1"] = "Priority"
    ws["C1"] = "Score"

    # Data row
    ws["A2"] = "Active"
    ws["B2"] = "High"
    ws["C2"] = 50

    # List validation (dropdown)
    dv_list = DataValidation(
        type="list",
        formula1='"Active,Inactive,Pending"',
        allow_blank=True,
    )
    dv_list.prompt = "Select a status"
    dv_list.error = "Invalid status"
    dv_list.add("A2:A10")
    ws.add_data_validation(dv_list)

    # Whole number validation (range 1-100)
    dv_whole = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="100",
        allow_blank=False,
    )
    dv_whole.prompt = "Enter a score between 1 and 100"
    dv_whole.error = "Score must be 1-100"
    dv_whole.add("C2:C10")
    ws.add_data_validation(dv_whole)

    wb.save(FIXTURES_DIR / "data_validations.xlsx")


def _create_with_images() -> None:
    """Create a fixture with images."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.title = "Sheet1"

    ws["A1"] = "Document with Image"

    # Create a simple 1x1 pixel PNG image for testing
    # PNG header + minimal IHDR + IDAT + IEND
    png_data = (
        b"\x89PNG\r\n\x1a\n"  # PNG signature
        b"\x00\x00\x00\rIHDR"  # IHDR chunk
        b"\x00\x00\x00\x01"  # width: 1
        b"\x00\x00\x00\x01"  # height: 1
        b"\x08\x02"  # bit depth 8, color type RGB
        b"\x00\x00\x00"  # compression, filter, interlace
        b"\x90wS\xde"  # CRC
        b"\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00\x00\x01\x01\x00\x05\x18\xd8N"
        b"\x00\x00\x00\x00IEND\xaeB`\x82"  # IEND chunk
    )

    # Save the test image
    test_image_path = FIXTURES_DIR / "test_image.png"
    test_image_path.write_bytes(png_data)

    # Add image to worksheet
    img = Image(str(test_image_path))
    img.width = 100
    img.height = 100
    ws.add_image(img, "B2")

    wb.save(FIXTURES_DIR / "with_images.xlsx")
