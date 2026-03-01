"""Cell extraction utilities.

This module provides functions to extract cell values and styles from
openpyxl worksheets.
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.utils import get_column_letter

from xldump.models import (
    AlignmentInfo,
    BorderInfo,
    BorderSideInfo,
    CellDump,
    CellStyle,
    FillInfo,
    FontInfo,
    ProtectionInfo,
)

if TYPE_CHECKING:
    from openpyxl.cell import Cell
    from openpyxl.styles import Alignment, Border, Fill, Font, Protection, Side
    from openpyxl.styles.colors import Color
    from openpyxl.worksheet.worksheet import Worksheet


def extract_cells(
    ws: Worksheet,
    *,
    include_styles: bool = True,
    include_values: bool = True,
) -> dict[str, dict[str, CellDump | None]]:
    """Extract all cells from a worksheet as a nested dict (row -> column).

    Args:
        ws: The openpyxl Worksheet to extract from.
        include_styles: Whether to include style information.
        include_values: Whether to include cell values.

    Returns:
        A dict mapping row numbers (as strings) to dicts mapping
        column letters to CellDump objects. Merged cell ranges have
        their top-left cell with data, others are None.

    """
    rows: dict[str, dict[str, CellDump | None]] = {}

    # Get merged cell ranges for special handling
    merged_ranges = list(ws.merged_cells.ranges)

    # Build a set of cells that are part of merged ranges but not the top-left
    merged_non_topleft: set[str] = set()
    for merged_range in merged_ranges:
        min_col = merged_range.min_col
        min_row = merged_range.min_row
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if row != min_row or col != min_col:
                    merged_non_topleft.add(f"{get_column_letter(col)}{row}")

    # Iterate through all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            if cell.coordinate is None:
                continue

            row_num = str(cell.row)
            col_letter = get_column_letter(cell.column)

            # Skip empty cells (no value and default style) unless they are merged
            if (
                cell.value is None
                and _is_default_style(cell)
                and cell.coordinate not in merged_non_topleft
            ):
                continue

            # Initialize row dict if needed
            if row_num not in rows:
                rows[row_num] = {}

            # For merged cells (not top-left), store None
            if cell.coordinate in merged_non_topleft:
                rows[row_num][col_letter] = None
                continue

            # Extract cell data
            cell_dump = _cell_to_dump(
                cell,
                include_styles=include_styles,
                include_values=include_values,
            )
            rows[row_num][col_letter] = cell_dump

    return rows


def _is_default_style(cell: Cell) -> bool:
    """Check if a cell has default styling (no custom styles applied)."""
    # Check font
    font = cell.font
    if font and (font.bold or font.italic or font.name != "Calibri" or font.size != 11):
        return False

    # Check fill
    fill = cell.fill
    if fill and fill.fill_type and fill.fill_type != "none":
        return False

    # Check alignment
    alignment = cell.alignment
    if alignment and (
        alignment.horizontal or alignment.vertical or alignment.wrap_text
    ):
        return False

    # Check number format
    return not (cell.number_format and cell.number_format != "General")


def _cell_to_dump(
    cell: Cell,
    *,
    include_styles: bool = True,
    include_values: bool = True,
) -> CellDump:
    """Convert an openpyxl Cell to a CellDump dataclass.

    Args:
        cell: The openpyxl Cell to convert.
        include_styles: Whether to include style information.
        include_values: Whether to include cell values.

    Returns:
        A CellDump object representing the cell.

    """
    value = cell.value if include_values else None
    data_type = cell.data_type if cell.data_type else "n"

    style: CellStyle | None = None
    if include_styles:
        style = CellStyle(
            font=_extract_font(cell.font) if cell.font else None,
            fill=_extract_fill(cell.fill) if cell.fill else None,
            alignment=_extract_alignment(cell.alignment) if cell.alignment else None,
            border=_extract_border(cell.border) if cell.border else None,
            number_format=cell.number_format if cell.number_format else "General",
            protection=_extract_protection(cell.protection)
            if cell.protection
            else None,
        )

    return CellDump(
        coordinate=cell.coordinate,
        value=value,
        data_type=data_type,
        style=style,
    )


def _extract_font(font: Font) -> FontInfo:
    """Extract font information from an openpyxl Font object."""
    return FontInfo(
        name=font.name,
        size=font.size,
        bold=bool(font.bold),
        italic=bool(font.italic),
        color=_resolve_color(font.color) if font.color else None,
    )


def _extract_fill(fill: Fill) -> FillInfo:
    """Extract fill information from an openpyxl Fill object."""
    fill_type = getattr(fill, "fill_type", None) or getattr(fill, "patternType", None)
    fg_color = None

    # Handle PatternFill
    if hasattr(fill, "fgColor") and fill.fgColor:
        fg_color = _resolve_color(fill.fgColor)
    elif hasattr(fill, "start_color") and fill.start_color:
        fg_color = _resolve_color(fill.start_color)

    return FillInfo(
        type=fill_type,
        fg_color=fg_color,
    )


def _extract_alignment(alignment: Alignment) -> AlignmentInfo:
    """Extract alignment information from an openpyxl Alignment object."""
    return AlignmentInfo(
        horizontal=alignment.horizontal,
        vertical=alignment.vertical,
        wrap_text=bool(alignment.wrap_text),
    )


def _extract_border(border: Border) -> BorderInfo:
    """Extract border information from an openpyxl Border object."""
    return BorderInfo(
        top=_extract_border_side(border.top) if border.top else None,
        bottom=_extract_border_side(border.bottom) if border.bottom else None,
        left=_extract_border_side(border.left) if border.left else None,
        right=_extract_border_side(border.right) if border.right else None,
    )


def _extract_border_side(side: Side) -> BorderSideInfo | None:
    """Extract border side information from an openpyxl Side object."""
    if not side.style:
        return None
    return BorderSideInfo(
        style=side.style,
        color=_resolve_color(side.color) if side.color else None,
    )


def _extract_protection(protection: Protection) -> ProtectionInfo:
    """Extract protection information from an openpyxl Protection object."""
    return ProtectionInfo(
        locked=bool(protection.locked) if protection.locked is not None else True,
        hidden=bool(protection.hidden) if protection.hidden is not None else False,
    )


def _resolve_color(color: Color | None) -> str | None:
    """Resolve an openpyxl Color to an ARGB hex string.

    Handles rgb, theme, and indexed color types.

    Args:
        color: The openpyxl Color object to resolve.

    Returns:
        The ARGB hex string (e.g., "FF000000") or None if unresolvable.

    """
    if color is None:
        return None

    # Handle direct RGB color
    if hasattr(color, "rgb") and color.rgb and color.rgb != "00000000":
        rgb = color.rgb
        # Handle both 6-digit and 8-digit hex
        if isinstance(rgb, str):
            if len(rgb) == 6:
                return f"FF{rgb}"
            return rgb

    # Handle indexed colors (simplified - just return None for now)
    # Full implementation would use the color palette
    if hasattr(color, "indexed") and color.indexed is not None:
        # Common indexed colors
        indexed_colors = {
            0: "FF000000",  # Black
            1: "FFFFFFFF",  # White
            2: "FFFF0000",  # Red
            3: "FF00FF00",  # Green
            4: "FF0000FF",  # Blue
            5: "FFFFFF00",  # Yellow
            6: "FFFF00FF",  # Magenta
            7: "FF00FFFF",  # Cyan
        }
        return indexed_colors.get(color.indexed)

    # Theme colors require workbook context (simplified for now)
    if hasattr(color, "theme") and color.theme is not None:
        # Basic theme color mapping (would need workbook.theme for full support)
        return None

    return None
