"""Sheet extraction utilities.

This module provides functions to extract sheet-level information from
openpyxl worksheets.
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from xldump.models import SheetInfo

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


def extract_sheet_info(ws: Worksheet, index: int) -> SheetInfo:
    """Extract lightweight sheet information for scan operation.

    Args:
        ws: The openpyxl Worksheet to extract from.
        index: The zero-based index of the sheet in the workbook.

    Returns:
        A SheetInfo object with summary information about the sheet.

    """
    # Get merged cell count (not available in read_only mode)
    merged_cell_count = 0
    if hasattr(ws, "merged_cells") and ws.merged_cells:
        merged_cell_count = len(list(ws.merged_cells.ranges))

    # Check for images (not available in read_only mode)
    has_images = False
    if hasattr(ws, "_images") and ws._images:
        has_images = len(ws._images) > 0

    # Check for data validations (not available in read_only mode)
    has_data_validations = False
    if hasattr(ws, "data_validations") and ws.data_validations:
        has_data_validations = len(ws.data_validations.dataValidation) > 0

    return SheetInfo(
        name=ws.title,
        index=index,
        dimensions=ws.dimensions if ws.dimensions else None,
        max_row=ws.max_row if ws.max_row else 0,
        max_column=ws.max_column if ws.max_column else 0,
        merged_cell_count=merged_cell_count,
        has_images=has_images,
        has_data_validations=has_data_validations,
    )
