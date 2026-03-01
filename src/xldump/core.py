"""Core functionality for xldump.

This module provides the main API functions for scanning and dumping
Excel workbook structures.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from xldump.extractors import (
    extract_cells,
    extract_images,
    extract_merged_cells,
    extract_sheet_info,
    extract_validations,
)
from xldump.models import ScanResult, SheetDump, SheetInfo, WorkbookDump


class XlDumpError(Exception):
    """Base exception for xldump errors."""

    pass


class FileNotFoundError(XlDumpError):
    """Raised when the input file is not found."""

    pass


class InvalidFileError(XlDumpError):
    """Raised when the input file is not a valid .xlsx file."""

    pass


def scan(filepath: str | Path) -> ScanResult:
    """Scan a workbook and return lightweight sheet information.

    This function reads sheet metadata including merged cell counts
    and other structural information.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        A ScanResult containing sheet names and summary information.

    Raises:
        FileNotFoundError: If the file does not exist.
        InvalidFileError: If the file is not a valid .xlsx file.

    """
    filepath = Path(filepath)
    _validate_file(filepath)

    # Note: We use read_only=False to access merged_cells and other metadata.
    # For very large files, consider using read_only=True with limited info.
    wb = load_workbook(filepath, read_only=False, data_only=False)
    try:
        sheets: list[SheetInfo] = []
        for index, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            sheet_info = extract_sheet_info(ws, index)
            sheets.append(sheet_info)

        return ScanResult(
            file=filepath.name,
            sheet_count=len(sheets),
            sheets=sheets,
        )
    finally:
        wb.close()


def dump(
    filepath: str | Path,
    *,
    sheets: list[str] | None = None,
    include_styles: bool = True,
    include_values: bool = True,
    include_merges: bool = True,
    include_borders: bool = True,
    include_validations: bool = True,
    include_images: bool = True,
    data_only: bool = False,
) -> WorkbookDump:
    """Dump the physical structure of a workbook to a WorkbookDump object.

    Args:
        filepath: Path to the .xlsx file.
        sheets: List of sheet names to dump. None means all sheets.
        include_styles: Whether to include style information.
        include_values: Whether to include cell values (False for style-only mode).
        include_merges: Whether to include merged cell information.
        include_borders: Whether to include border information (part of styles).
        include_validations: Whether to include data validation rules.
        include_images: Whether to include image position information.
        data_only: If True, use cached formula results instead of formulas.

    Returns:
        A WorkbookDump containing all requested sheet information.

    Raises:
        FileNotFoundError: If the file does not exist.
        InvalidFileError: If the file is not a valid .xlsx file.

    """
    filepath = Path(filepath)
    _validate_file(filepath)

    wb = load_workbook(filepath, read_only=False, data_only=data_only)
    try:
        sheet_dumps: list[SheetDump] = []
        target_sheets = sheets if sheets else wb.sheetnames

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            index = wb.sheetnames.index(sheet_name)
            sheet_dump = _dump_sheet(
                ws,
                wb,
                index,
                include_styles=include_styles,
                include_values=include_values,
                include_merges=include_merges,
                include_validations=include_validations,
                include_images=include_images,
            )
            sheet_dumps.append(sheet_dump)

        return WorkbookDump(
            file=filepath.name,
            sheets=sheet_dumps,
            sheet_count=len(sheet_dumps),
        )
    finally:
        wb.close()


def dump_sheet(
    filepath: str | Path,
    sheet_name: str,
    *,
    include_styles: bool = True,
    include_values: bool = True,
    include_merges: bool = True,
    include_borders: bool = True,
    include_validations: bool = True,
    include_images: bool = True,
    data_only: bool = False,
) -> SheetDump:
    """Dump a single sheet's physical structure.

    This is a convenience function for dumping just one sheet.

    Args:
        filepath: Path to the .xlsx file.
        sheet_name: Name of the sheet to dump.
        include_styles: Whether to include style information.
        include_values: Whether to include cell values.
        include_merges: Whether to include merged cell information.
        include_borders: Whether to include border information.
        include_validations: Whether to include data validation rules.
        include_images: Whether to include image position information.
        data_only: If True, use cached formula results instead of formulas.

    Returns:
        A SheetDump containing the sheet's structure.

    Raises:
        FileNotFoundError: If the file does not exist.
        InvalidFileError: If the file is not a valid .xlsx file.
        ValueError: If the sheet name is not found.

    """
    filepath = Path(filepath)
    _validate_file(filepath)

    wb = load_workbook(filepath, read_only=False, data_only=data_only)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws = wb[sheet_name]
        index = wb.sheetnames.index(sheet_name)

        return _dump_sheet(
            ws,
            wb,
            index,
            include_styles=include_styles,
            include_values=include_values,
            include_merges=include_merges,
            include_validations=include_validations,
            include_images=include_images,
        )
    finally:
        wb.close()


def _validate_file(filepath: Path) -> None:
    """Validate that the file exists and has a .xlsx extension.

    Args:
        filepath: Path to validate.

    Raises:
        FileNotFoundError: If the file does not exist.
        InvalidFileError: If the file is not a .xlsx file.

    """
    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    if filepath.suffix.lower() not in (".xlsx", ".xlsm"):
        raise InvalidFileError(
            f"Invalid file type: {filepath.suffix}. "
            "Only .xlsx and .xlsm files are supported. "
            "Note: .xls (old Excel format) is not supported."
        )


def _dump_sheet(
    ws,  # Worksheet type
    wb,  # Workbook type for theme color resolution
    index: int,
    *,
    include_styles: bool,
    include_values: bool,
    include_merges: bool,
    include_validations: bool,
    include_images: bool,
) -> SheetDump:
    """Dump a single worksheet to a SheetDump object.

    Args:
        ws: The openpyxl Worksheet to dump.
        wb: The openpyxl Workbook for theme color resolution.
        index: The zero-based index of the sheet.
        include_styles: Whether to include style information.
        include_values: Whether to include cell values.
        include_merges: Whether to include merged cell information.
        include_validations: Whether to include data validation rules.
        include_images: Whether to include image position information.

    Returns:
        A SheetDump containing the sheet's structure.

    """
    # Extract cells (with workbook for theme color resolution)
    rows = extract_cells(
        ws,
        include_styles=include_styles,
        include_values=include_values,
        workbook=wb,
    )

    # Extract merged cells
    merged_cells = []
    if include_merges:
        merged_cells = extract_merged_cells(ws)

    # Extract data validations
    data_validations = []
    if include_validations:
        data_validations = extract_validations(ws)

    # Extract images
    images = []
    if include_images:
        images = extract_images(ws)

    return SheetDump(
        name=ws.title,
        index=index,
        dimensions=ws.dimensions if ws.dimensions else None,
        max_row=ws.max_row if ws.max_row else 0,
        max_column=ws.max_column if ws.max_column else 0,
        merged_cells=merged_cells,
        data_validations=data_validations,
        images=images,
        rows=rows,
    )
